import os
import flet as ft
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, PatternFill, Font


# =========================================================
# ☑️ 공통 스타일
# =========================================================
BORDER_COLOR = "#5A5A5A"
HEADER_BG = "#F5F5F5"
WHITE = "#FFFFFF"
TEXT_COLOR = "#111111"
PAGE_BG = "#EDEDED"

DIALOG_WIDTH = 1500
DIALOG_HEIGHT = 900

FIELD_HEIGHT = 42
ROW_HEIGHT = 42
TITLE_HEIGHT = 150

# =========================================================
# ☑️ 문서 전체 공통 폭
# =========================================================
DOC_WIDTH = 1424

# 상단 영역 폭
TITLE_LEFT_WIDTH = 1114
TITLE_RIGHT_WIDTH = 310

APPROVAL_COL_WIDTH = 103

INFO_LABEL_W = 120
INFO_LEFT_VALUE_W = 730
INFO_RIGHT_LABEL_W = 120
INFO_RIGHT_VALUE_W = DOC_WIDTH - INFO_LABEL_W - INFO_LEFT_VALUE_W - INFO_RIGHT_LABEL_W  # 454

# =========================================================
# ☑️ 주문처 / TEL / FAX / 담당자 블록 폭
# =========================================================
CUSTOMER_LABEL_W = 120
MID_LABEL_W = 140
MID_VALUE_W = 260

RIGHT_LABEL_W = 130
RIGHT_INPUT_W = 150
RIGHT_TOTAL_W = RIGHT_LABEL_W + RIGHT_INPUT_W

CUSTOMER_VALUE_W = DOC_WIDTH - CUSTOMER_LABEL_W - MID_LABEL_W - MID_VALUE_W - RIGHT_TOTAL_W
CUSTOMER_BLOCK_HEIGHT = 84

# =========================================================
# ☑️ 하단 품목 폭
# =========================================================
COL_NO = 50
COL_LOT = 110

# 🔥 수정: 품명이 너무 좁아서 120 → 240으로 확대
COL_NAME = 240

COL_SPEC = 120
COL_UNIT = 70
COL_BUY = 110
COL_SELL = 110
COL_QTY = 90
COL_BUY_SUM = 120
COL_SELL_SUM = 120

# 🔥 수정: 생산기간은 남는 폭을 가져가되, 품명 폭을 키운 만큼 자동으로 줄어듦
COL_PERIOD = 120

ITEM_TOTAL_WIDTH = (
    COL_NO + COL_LOT + COL_NAME + COL_SPEC + COL_UNIT +
    COL_BUY + COL_SELL + COL_QTY + COL_BUY_SUM + COL_SELL_SUM + COL_PERIOD
)

COL_PERIOD = COL_PERIOD + (DOC_WIDTH - ITEM_TOTAL_WIDTH)


# =========================================================
# ☑️ 숫자 변환 유틸
# =========================================================
def to_int(value):
    # ⭐ 수량/구매단가/판매가 계산 전에 문자열 숫자를 정수로 바꾸는 공통 유틸
    try:
        value = str(value).replace(",", "").strip()
        if value == "":
            return 0
        return int(value)
    except:
        return 0


def format_number(value):
    # ⭐ 계산된 구매가계/판매가계를 10,000 같은 형식으로 표시
    if value in ("", None):
        return ""
    try:
        return f"{int(value):,}"
    except:
        return str(value)


# =========================================================
# ☑️ 엑셀 스타일 공통 유틸
# =========================================================
# ❌ 엑셀 셀 범위 스타일 공통 함수
# ❌ 테두리 / 정렬 / 폰트 / 배경색 / 값 입력을 한 번에 처리함
# ❌ build_excel_workbook() 안에서 반복 호출되며 엑셀 표 모양을 유지하는 핵심 함수
def apply_range_style(
    ws,
    cell_range,
    value="",
    fill_color=None,
    bold=False,
    font_size=12,
    horizontal="center",
    vertical="center",
):
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    cells = ws[cell_range]
    for row in cells:
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(
                horizontal=horizontal,
                vertical=vertical,
                wrap_text=False,
                shrink_to_fit=True,
            )
            cell.font = Font(
                bold=bold,
                size=font_size,
                name="맑은 고딕",
            )
            if fill_color:
                cell.fill = PatternFill(
                    fill_type="solid",
                    fgColor=fill_color,
                )

    top_left = cells[0][0]
    top_left.value = value


# ❌ 지정한 범위 전체에 테두리와 기본 정렬/폰트를 넣는 함수
# ❌ 품목 표 전체 줄을 그릴 때 사용됨
def set_all_borders(ws, start_row, end_row, start_col, end_col):
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            ws.cell(r, c).border = border
            ws.cell(r, c).alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=False,
                shrink_to_fit=True,
            )
            ws.cell(r, c).font = Font(name="맑은 고딕", size=11)


# =========================================================
# ☑️ 공통 셀 박스
# =========================================================
def box_label(text, width=None, height=FIELD_HEIGHT, bold=False):
    return ft.Container(
        width=width,
        height=height,
        bgcolor=HEADER_BG,
        alignment=ft.Alignment(0, 0),
        border=ft.Border.all(1, BORDER_COLOR),
        content=ft.Text(
            text,
            size=13,
            color=TEXT_COLOR,
            weight=ft.FontWeight.W_700 if bold else ft.FontWeight.W_500,
            text_align=ft.TextAlign.CENTER,
        ),
    )


def box_input(control, width=None, height=FIELD_HEIGHT):
    return ft.Container(
        width=width,
        height=height,
        border=ft.Border.all(1, BORDER_COLOR),
        bgcolor=WHITE,
        alignment=ft.Alignment(0, 0),
        content=control,
    )


def build_textfield(
    value="",
    text_align=ft.TextAlign.LEFT,
    read_only=False,
    on_change=None,
    dense=True,
):
    return ft.TextField(
        value=value,
        border=ft.InputBorder.NONE,
        filled=False,
        text_size=13,
        cursor_height=18,
        content_padding=ft.Padding(left=8, right=8, top=10, bottom=10),
        text_align=text_align,
        read_only=read_only,
        on_change=on_change,
        dense=dense,
    )


# =========================================================
# ☑️ 생산지시서 모달 클래스
# =========================================================
class ProductionOrderDialog:
    def __init__(self, page: ft.Page):
        self.page = page

        # -------------------------------------------------
        # ☑️ 상단 기본 정보 컨트롤
        # -------------------------------------------------
        self.instruction_date = build_textfield()
        self.doc_no = build_textfield()
        self.contract_no = build_textfield()

        self.delivery_place = build_textfield()
        self.request_dept = build_textfield()
        self.production_manager = build_textfield()

        self.customer_name = build_textfield()
        self.tel = build_textfield()
        self.fax = build_textfield()
        self.person_in_charge = build_textfield()

        self.approval_1 = build_textfield(text_align=ft.TextAlign.CENTER)
        self.approval_2 = build_textfield(text_align=ft.TextAlign.CENTER)
        self.approval_3 = build_textfield(text_align=ft.TextAlign.CENTER)

        # -------------------------------------------------
        # ☑️ 하단 품목 행 데이터
        # -------------------------------------------------
        self.item_rows = []
        for i in range(18):
            self.item_rows.append(self.build_item_row(i + 1))

        # -------------------------------------------------
        # ☑️ 다이얼로그 생성
        # -------------------------------------------------
        self.dialog = ft.AlertDialog(
            modal=True,
            inset_padding=10,
            bgcolor=ft.Colors.TRANSPARENT,
            content_padding=0,
            shape=ft.RoundedRectangleBorder(radius=12),
            content=self.build_dialog_content(),
        )

    # =====================================================
    # ☑️ 품목 한 줄 생성
    # =====================================================
    def build_item_row(self, no: int):
        row_data = {}

        def on_amount_change(e=None):
            qty = to_int(row_data["qty"].value)
            buy_price = to_int(row_data["buy_price"].value)
            sell_price = to_int(row_data["sell_price"].value)

            buy_total = qty * buy_price
            sell_total = qty * sell_price

            row_data["buy_total"].value = format_number(buy_total) if buy_total else ""
            row_data["sell_total"].value = format_number(sell_total) if sell_total else ""
            self.page.update()

        row_data["no"] = no
        row_data["lot_no"] = build_textfield(text_align=ft.TextAlign.CENTER)
        row_data["product_name"] = build_textfield()
        row_data["spec"] = build_textfield()
        row_data["unit"] = build_textfield(text_align=ft.TextAlign.CENTER)
        row_data["buy_price"] = build_textfield(
            text_align=ft.TextAlign.RIGHT,
            on_change=on_amount_change,
        )
        row_data["sell_price"] = build_textfield(
            text_align=ft.TextAlign.RIGHT,
            on_change=on_amount_change,
        )
        row_data["qty"] = build_textfield(
            text_align=ft.TextAlign.RIGHT,
            on_change=on_amount_change,
        )
        row_data["buy_total"] = build_textfield(
            text_align=ft.TextAlign.RIGHT,
            read_only=True,
        )
        row_data["sell_total"] = build_textfield(
            text_align=ft.TextAlign.RIGHT,
            read_only=True,
        )
        row_data["period"] = build_textfield(text_align=ft.TextAlign.CENTER)

        return row_data

    # =====================================================
    # ☑️ 상단 제목 영역
    # =====================================================
    def build_title_area(self):
        return ft.Container(
            width=DOC_WIDTH,
            content=ft.Row(
                spacing=0,
                controls=[
                    ft.Container(
                        width=TITLE_LEFT_WIDTH,
                        height=TITLE_HEIGHT,
                        border=ft.Border.all(1, BORDER_COLOR),
                        alignment=ft.Alignment(0, 0),
                        content=ft.Text(
                            "생산지시서",
                            size=28,
                            weight=ft.FontWeight.W_700,
                            color=TEXT_COLOR,
                        ),
                    ),
                    ft.Container(
                        width=TITLE_RIGHT_WIDTH,
                        height=TITLE_HEIGHT,
                        content=ft.Column(
                            spacing=0,
                            controls=[
                                ft.Row(
                                    spacing=0,
                                    controls=[
                                        box_label("담당", width=APPROVAL_COL_WIDTH),
                                        box_input(self.approval_1, width=APPROVAL_COL_WIDTH),
                                        box_input(self.approval_2, width=APPROVAL_COL_WIDTH + 1),
                                    ],
                                ),
                                ft.Row(
                                    spacing=0,
                                    controls=[
                                        box_input(build_textfield(), width=APPROVAL_COL_WIDTH, height=108),
                                        box_input(build_textfield(), width=APPROVAL_COL_WIDTH, height=108),
                                        box_input(build_textfield(), width=APPROVAL_COL_WIDTH + 1, height=108),
                                    ],
                                ),
                            ],
                        ),
                    ),
                ],
            ),
        )

    # =====================================================
    # ☑️ 상단 기본 정보 행
    # =====================================================
    def build_info_row(self, left_title, left_control, right_title, right_control):
        return ft.Container(
            width=DOC_WIDTH,
            content=ft.Row(
                spacing=0,
                controls=[
                    box_label(left_title, width=INFO_LABEL_W),
                    box_input(left_control, width=INFO_LEFT_VALUE_W),
                    box_label(right_title, width=INFO_RIGHT_LABEL_W),
                    box_input(right_control, width=INFO_RIGHT_VALUE_W),
                ],
            ),
        )

    # =====================================================
    # ☑️ 담당자 블록
    # =====================================================
    def build_person_block(self):
        return ft.Container(
            width=RIGHT_TOTAL_W,
            height=CUSTOMER_BLOCK_HEIGHT,
            border=ft.Border.all(1, BORDER_COLOR),
            bgcolor=WHITE,
            clip_behavior=ft.ClipBehavior.HARD_EDGE,
            content=ft.Row(
                spacing=0,
                controls=[
                    ft.Container(
                        width=RIGHT_LABEL_W,
                        height=CUSTOMER_BLOCK_HEIGHT,
                        bgcolor=HEADER_BG,
                        alignment=ft.Alignment(0, 0),
                        border=ft.Border.only(
                            right=ft.BorderSide(1, BORDER_COLOR),
                        ),
                        content=ft.Text(
                            "담당자",
                            size=13,
                            color=TEXT_COLOR,
                            weight=ft.FontWeight.W_500,
                            text_align=ft.TextAlign.CENTER,
                        ),
                    ),
                    ft.Container(
                        expand=True,
                        height=CUSTOMER_BLOCK_HEIGHT,
                        bgcolor=WHITE,
                        alignment=ft.Alignment(0, 0),
                        content=self.person_in_charge,
                    ),
                ],
            ),
        )

    # =====================================================
    # ☑️ 주문처 / TEL / FAX / 담당자 블록
    # =====================================================
    def build_customer_block(self):
        return ft.Container(
            width=DOC_WIDTH,
            content=ft.Row(
                spacing=0,
                controls=[
                    box_label("주문처", width=CUSTOMER_LABEL_W, height=CUSTOMER_BLOCK_HEIGHT),
                    box_input(self.customer_name, width=CUSTOMER_VALUE_W, height=CUSTOMER_BLOCK_HEIGHT),
                    ft.Column(
                        spacing=0,
                        controls=[
                            ft.Row(
                                spacing=0,
                                controls=[
                                    box_label("TEL", width=MID_LABEL_W, bold=True),
                                    box_input(self.tel, width=MID_VALUE_W),
                                ],
                            ),
                            ft.Row(
                                spacing=0,
                                controls=[
                                    box_label("FAX", width=MID_LABEL_W, bold=True),
                                    box_input(self.fax, width=MID_VALUE_W),
                                ],
                            ),
                        ],
                    ),
                    self.build_person_block(),
                ],
            ),
        )

    # =====================================================
    # ☑️ 하단 품목 헤더
    # =====================================================
    def build_item_header(self):
        return ft.Container(
            width=DOC_WIDTH,
            content=ft.Row(
                spacing=0,
                controls=[
                    box_label("no", width=COL_NO, height=44, bold=True),
                    box_label("LOT NO", width=COL_LOT, height=44, bold=True),
                    box_label("품명", width=COL_NAME, height=44, bold=True),
                    box_label("규격/사양", width=COL_SPEC, height=44, bold=True),
                    box_label("단위", width=COL_UNIT, height=44, bold=True),
                    box_label("구매단가", width=COL_BUY, height=44, bold=True),
                    box_label("판매가", width=COL_SELL, height=44, bold=True),
                    box_label("수량", width=COL_QTY, height=44, bold=True),
                    box_label("구매가계", width=COL_BUY_SUM, height=44, bold=True),
                    box_label("판매가계", width=COL_SELL_SUM, height=44, bold=True),
                    box_label("생산기간", width=COL_PERIOD, height=44, bold=True),
                ],
            ),
        )

    # =====================================================
    # ☑️ 하단 품목 한 줄 UI
    # =====================================================
    def build_item_line(self, row_data):
        return ft.Container(
            width=DOC_WIDTH,
            content=ft.Row(
                spacing=0,
                controls=[
                    box_input(
                        build_textfield(
                            value=str(row_data["no"]),
                            text_align=ft.TextAlign.CENTER,
                            read_only=True,
                        ),
                        width=COL_NO,
                        height=ROW_HEIGHT,
                    ),
                    box_input(row_data["lot_no"], width=COL_LOT, height=ROW_HEIGHT),
                    box_input(row_data["product_name"], width=COL_NAME, height=ROW_HEIGHT),
                    box_input(row_data["spec"], width=COL_SPEC, height=ROW_HEIGHT),
                    box_input(row_data["unit"], width=COL_UNIT, height=ROW_HEIGHT),
                    box_input(row_data["buy_price"], width=COL_BUY, height=ROW_HEIGHT),
                    box_input(row_data["sell_price"], width=COL_SELL, height=ROW_HEIGHT),
                    box_input(row_data["qty"], width=COL_QTY, height=ROW_HEIGHT),
                    box_input(row_data["buy_total"], width=COL_BUY_SUM, height=ROW_HEIGHT),
                    box_input(row_data["sell_total"], width=COL_SELL_SUM, height=ROW_HEIGHT),
                    box_input(row_data["period"], width=COL_PERIOD, height=ROW_HEIGHT),
                ],
            ),
        )

    # =====================================================
    # ☑️ 품목 입력 영역
    # =====================================================
    def build_items_section(self):
        return ft.Container(
            width=DOC_WIDTH,
            content=ft.Column(
                spacing=0,
                controls=[self.build_item_header()]
                + [self.build_item_line(row_data) for row_data in self.item_rows],
            ),
        )

    # =====================================================
    # ☑️ 저장 데이터 수집
    # =====================================================
    # ❌ 화면 입력값들을 dict 형태의 저장 데이터로 모으는 핵심 함수
    # ❌ 엑셀 생성 전에 build_excel_workbook()가 이 데이터를 가져다 씀
    def collect_data(self):
        items = []

        for row in self.item_rows:
            item = {
                "no": row["no"],
                "lot_no": row["lot_no"].value,
                "product_name": row["product_name"].value,
                "spec": row["spec"].value,
                "unit": row["unit"].value,
                "buy_price": row["buy_price"].value,
                "sell_price": row["sell_price"].value,
                "qty": row["qty"].value,
                "buy_total": row["buy_total"].value,
                "sell_total": row["sell_total"].value,
                "period": row["period"].value,
            }

            has_value = any(str(v).strip() != "" for k, v in item.items() if k != "no")
            if has_value:
                items.append(item)

        return {
            "instruction_date": self.instruction_date.value,
            "doc_no": self.doc_no.value,
            "contract_no": self.contract_no.value,
            "delivery_place": self.delivery_place.value,
            "request_dept": self.request_dept.value,
            "production_manager": self.production_manager.value,
            "customer_name": self.customer_name.value,
            "tel": self.tel.value,
            "fax": self.fax.value,
            "person_in_charge": self.person_in_charge.value,
            "items": items,
        }

    # ❌ 생산지시서 화면 구조를 엑셀 워크북으로 그대로 다시 만드는 핵심 함수
    # ❌ 컬럼 폭 / 행 높이 / 병합 / 제목 / 상단정보 / 주문처블록 / 품목표 / 시트옵션이 전부 여기 있음
    def build_excel_workbook(self):
        data = self.collect_data()

        wb = Workbook()
        ws = wb.active
        ws.title = "생산지시서"

        # -------------------------------------------------
        # ☑️ 컬럼 폭 설정
        # -------------------------------------------------
        column_widths = {
            "A": 12,
            "B": 16,

            # 🔥 수정: 엑셀에서도 품명 칸 확대
            "C": 28,

            "D": 16,
            "E": 12,
            "F": 14,
            "G": 14,
            "H": 12,
            "I": 14,
            "J": 16,

            # 🔥 수정: 엑셀에서도 생산기간 칸 축소
            "K": 12,
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # -------------------------------------------------
        # ☑️ 행 높이 설정
        # -------------------------------------------------
        for row_idx in range(1, 60):
            ws.row_dimensions[row_idx].height = 26

        ws.row_dimensions[1].height = 30
        ws.row_dimensions[2].height = 30
        ws.row_dimensions[3].height = 30
        ws.row_dimensions[4].height = 30
        ws.row_dimensions[5].height = 26
        ws.row_dimensions[6].height = 26
        ws.row_dimensions[7].height = 26
        ws.row_dimensions[8].height = 26
        ws.row_dimensions[9].height = 26
        ws.row_dimensions[10].height = 28

        # -------------------------------------------------
        # ☑️ 머지
        # -------------------------------------------------
        ws.merge_cells("A1:H4")

        # 🔥 수정: 오른쪽 결재 영역은 통병합 금지
        ws.merge_cells("I2:I4")
        ws.merge_cells("J2:J4")
        ws.merge_cells("K2:K4")

        ws.merge_cells("B5:H5")
        ws.merge_cells("J5:K5")

        ws.merge_cells("B6:H6")
        ws.merge_cells("J6:K6")

        ws.merge_cells("B7:H7")
        ws.merge_cells("J7:K7")

        ws.merge_cells("A8:A9")
        ws.merge_cells("B8:E9")
        ws.merge_cells("G8:H8")
        ws.merge_cells("G9:H9")
        ws.merge_cells("I8:I9")
        ws.merge_cells("J8:K9")

        # -------------------------------------------------
        # ☑️ 공통 색상
        # -------------------------------------------------
        header_fill = "F2F2F2"

        # -------------------------------------------------
        # ☑️ 제목 영역
        # -------------------------------------------------
        apply_range_style(
            ws,
            "A1:H4",
            value="생산지시서",
            bold=True,
            font_size=24,
        )

        # 🔥 수정: 오른쪽 상단은 칸별로 따로 스타일 적용
        apply_range_style(ws, "I1:I1", value="담당", fill_color=header_fill, bold=True)
        apply_range_style(ws, "J1:J1", value="", fill_color=header_fill, bold=True)
        apply_range_style(ws, "K1:K1", value="", fill_color=header_fill, bold=True)

        apply_range_style(ws, "I2:I4", value="")
        apply_range_style(ws, "J2:J4", value="")
        apply_range_style(ws, "K2:K4", value="")

        # -------------------------------------------------
        # ☑️ 상단 기본 정보
        # -------------------------------------------------
        apply_range_style(ws, "A5:A5", value="지시일자", fill_color=header_fill, bold=True)
        apply_range_style(ws, "B5:H5", value=data["instruction_date"], horizontal="left")
        apply_range_style(ws, "I5:I5", value="납품장소", fill_color=header_fill, bold=True)
        apply_range_style(ws, "J5:K5", value=data["delivery_place"], horizontal="left")

        apply_range_style(ws, "A6:A6", value="문서번호", fill_color=header_fill, bold=True)
        apply_range_style(ws, "B6:H6", value=data["doc_no"], horizontal="left")
        apply_range_style(ws, "I6:I6", value="요구부서", fill_color=header_fill, bold=True)
        apply_range_style(ws, "J6:K6", value=data["request_dept"], horizontal="left")

        apply_range_style(ws, "A7:A7", value="계약번호", fill_color=header_fill, bold=True)
        apply_range_style(ws, "B7:H7", value=data["contract_no"], horizontal="left")
        apply_range_style(ws, "I7:I7", value="생산담당자", fill_color=header_fill, bold=True)
        apply_range_style(ws, "J7:K7", value=data["production_manager"], horizontal="left")

        # -------------------------------------------------
        # ☑️ 주문처 / TEL / FAX / 담당자 블록
        # -------------------------------------------------
        apply_range_style(ws, "A8:A9", value="주문처", fill_color=header_fill, bold=True)
        apply_range_style(ws, "B8:E9", value=data["customer_name"], horizontal="left")

        apply_range_style(ws, "F8:F8", value="TEL", fill_color=header_fill, bold=True)
        apply_range_style(ws, "G8:H8", value=data["tel"], horizontal="left")

        apply_range_style(ws, "F9:F9", value="FAX", fill_color=header_fill, bold=True)
        apply_range_style(ws, "G9:H9", value=data["fax"], horizontal="left")

        apply_range_style(ws, "I8:I9", value="담당자", fill_color=header_fill, bold=True)
        apply_range_style(ws, "J8:K9", value=data["person_in_charge"], horizontal="left")

        # -------------------------------------------------
        # ☑️ 품목 헤더
        # -------------------------------------------------
        item_header_row = 10
        headers = [
            "no",
            "LOT NO",
            "품명",
            "규격/사양",
            "단위",
            "구매단가",
            "판매가",
            "수량",
            "구매가계",
            "판매가계",
            "생산기간",
        ]

        for idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=item_header_row, column=idx)
            cell.value = header
            cell.fill = PatternFill(fill_type="solid", fgColor=header_fill)
            cell.font = Font(bold=True, size=11, name="맑은 고딕")
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=False,
                shrink_to_fit=True,
            )

        set_all_borders(ws, item_header_row, item_header_row, 1, 11)

        # -------------------------------------------------
        # ☑️ 품목 데이터
        # -------------------------------------------------
        start_row = 11
        max_item_rows = 18
        items = data["items"]

        if not items:
            for i in range(max_item_rows):
                row_no = start_row + i
                ws.cell(row=row_no, column=1).value = str(i + 1)
            set_all_borders(ws, start_row, start_row + max_item_rows - 1, 1, 11)
        else:
            for i in range(max_item_rows):
                row_no = start_row + i
                ws.cell(row=row_no, column=1).value = str(i + 1)

            for i, item in enumerate(items[:max_item_rows], start=0):
                row_no = start_row + i
                ws.cell(row=row_no, column=1).value = item["no"]
                ws.cell(row=row_no, column=2).value = item["lot_no"]
                ws.cell(row=row_no, column=3).value = item["product_name"]
                ws.cell(row=row_no, column=4).value = item["spec"]
                ws.cell(row=row_no, column=5).value = item["unit"]
                ws.cell(row=row_no, column=6).value = item["buy_price"]
                ws.cell(row=row_no, column=7).value = item["sell_price"]
                ws.cell(row=row_no, column=8).value = item["qty"]
                ws.cell(row=row_no, column=9).value = item["buy_total"]
                ws.cell(row=row_no, column=10).value = item["sell_total"]
                ws.cell(row=row_no, column=11).value = item["period"]

            set_all_borders(ws, start_row, start_row + max_item_rows - 1, 1, 11)

        # -------------------------------------------------
        # ☑️ 시트 옵션
        # -------------------------------------------------
        ws.freeze_panes = None
        ws.sheet_view.showGridLines = False

        return wb

    # =====================================================
    # ☑️ 엑셀 내보내기 버튼
    # =====================================================
    # ❌ build_excel_workbook()로 만든 워크북을 실제 .xlsx 파일로 저장하는 함수
    # ❌ 저장 경로 결정, wb.save(), 저장 완료 메시지 출력이 전부 여기서 처리됨
    def export_to_excel(self, e):
        try:
            wb = self.build_excel_workbook()

            # 🔥 핵심 수정 1: 상대경로 말고 현재 py 파일 기준 절대경로 사용
            base_dir = os.path.dirname(os.path.abspath(__file__))
            save_path = os.path.join(base_dir, "production_order.xlsx")

            # 🔥 핵심 수정 2: 실제 저장
            wb.save(save_path)

            # 🔥 핵심 수정 3: 콘솔에도 저장 위치 출력
            print("엑셀 저장 완료:", save_path)

            self.page.snack_bar = ft.SnackBar(
                content=ft.Text(f"엑셀 파일이 저장되었습니다.\n{save_path}"),
                open=True,
            )
            self.page.update()

        except Exception as ex:
            print("엑셀 저장 오류:", ex)

            self.page.snack_bar = ft.SnackBar(
                content=ft.Text(f"엑셀 저장 중 오류가 발생했습니다: {ex}"),
                open=True,
            )
            self.page.update()

    # =====================================================
    # ☑️ 저장 버튼
    # =====================================================
    def save_data(self, e):
        data = self.collect_data()

        print("===== 생산지시서 저장 데이터 =====")
        print(data)

        self.page.snack_bar = ft.SnackBar(
            content=ft.Text("생산지시서가 저장되었습니다."),
            open=True,
        )
        self.page.update()

    # =====================================================
    # ☑️ 닫기 버튼
    # =====================================================
    def close_dialog(self, e=None):
        self.page.pop_dialog()
        self.page.update()

    # =====================================================
    # ☑️ 다이얼로그 전체 콘텐츠
    # =====================================================
    def build_dialog_content(self):
        return ft.Container(
            width=DIALOG_WIDTH,
            height=DIALOG_HEIGHT,
            bgcolor=WHITE,
            border_radius=12,
            padding=20,
            content=ft.Column(
                spacing=0,
                controls=[
                    ft.Row(
                        alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
                        vertical_alignment=ft.CrossAxisAlignment.CENTER,
                        controls=[
                            ft.Text(
                                "생산지시서 입력",
                                size=20,
                                weight=ft.FontWeight.W_700,
                                color=TEXT_COLOR,
                            ),
                            ft.IconButton(
                                icon=ft.Icons.CLOSE,
                                on_click=self.close_dialog,
                            ),
                        ],
                    ),
                    ft.Container(height=16),
                    ft.Container(
                        expand=True,
                        content=ft.Column(
                            scroll=ft.ScrollMode.AUTO,
                            spacing=0,
                            horizontal_alignment=ft.CrossAxisAlignment.CENTER,
                            controls=[
                                self.build_title_area(),
                                self.build_info_row("지시일자", self.instruction_date, "납품장소", self.delivery_place),
                                self.build_info_row("문서번호", self.doc_no, "요구부서", self.request_dept),
                                self.build_info_row("계약번호", self.contract_no, "생산담당자", self.production_manager),
                                self.build_customer_block(),
                                self.build_items_section(),
                            ],
                        ),
                    ),
                    ft.Container(height=16),
                    ft.Row(
                        alignment=ft.MainAxisAlignment.END,
                        controls=[
                            ft.ElevatedButton(
                                "닫기",
                                height=42,
                                style=ft.ButtonStyle(
                                    shape=ft.RoundedRectangleBorder(radius=8),
                                ),
                                on_click=self.close_dialog,
                            ),
                            ft.ElevatedButton(
                                "엑셀 내보내기",
                                height=42,
                                style=ft.ButtonStyle(
                                    bgcolor="#16A34A",
                                    color=ft.Colors.WHITE,
                                    shape=ft.RoundedRectangleBorder(radius=8),
                                ),
                                on_click=self.export_to_excel,
                            ),
                            ft.ElevatedButton(
                                "저장",
                                height=42,
                                style=ft.ButtonStyle(
                                    bgcolor="#2563EB",
                                    color=ft.Colors.WHITE,
                                    shape=ft.RoundedRectangleBorder(radius=8),
                                ),
                                on_click=self.save_data,
                            ),
                        ],
                    ),
                ],
            ),
        )

    # =====================================================
    # ☑️ 모달 열기
    # =====================================================
    def open(self, e=None):
        self.page.show_dialog(self.dialog)
        self.page.update()


# =========================================================
# ☑️ 메인
# =========================================================
def main(page: ft.Page):
    page.title = "생산지시서 모달 예제"
    page.bgcolor = PAGE_BG
    page.padding = 30
    page.scroll = ft.ScrollMode.AUTO

    popup = ProductionOrderDialog(page)

    production_order = ft.Column(
        alignment=ft.MainAxisAlignment.CENTER,
        horizontal_alignment=ft.CrossAxisAlignment.CENTER,
        controls=[
            ft.Text(
                "생산지시서 모달창 예제",
                size=26,
                weight=ft.FontWeight.W_700,
            ),
            ft.Container(height=20),
            ft.ElevatedButton(
                "생산지시서 열기",
                width=220,
                height=50,
                style=ft.ButtonStyle(
                    bgcolor="#111827",
                    color=ft.Colors.WHITE,
                    shape=ft.RoundedRectangleBorder(radius=10),
                ),
                on_click=popup.open,
            ),
        ],
    )

    page.add(production_order)


if __name__ == "__main__":
    ft.run(main)