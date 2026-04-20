import os
import flet as ft
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, PatternFill, Font


# =========================================================
# ☑️ 공통 스타일
# =========================================================
BORDER_COLOR = "#5A5A5A"
HEADER_BG = "#F5F5F5"
WHITE = "#FFFFFF"
TEXT_COLOR = "#111111"
PAGE_BG = "#EDEDED"

DIALOG_WIDTH = 1500
DIALOG_HEIGHT = 920

FIELD_HEIGHT = 42
ROW_HEIGHT = 42
TITLE_HEIGHT = 150

# =========================================================
# ☑️ 문서 전체 공통 폭
# =========================================================
DOC_WIDTH = 1424

# 상단 제목 / 결재 영역
TITLE_LEFT_WIDTH = 1114
TITLE_RIGHT_WIDTH = 310
APPROVAL_COL_WIDTH = 103

# =========================================================
# ☑️ 수신처 블록 폭
# =========================================================
RECEIVER_LABEL_W = 120
RECEIVER_NAME_W = 460

RECEIVER_MID_LABEL_W = 120
RECEIVER_MID_VALUE_W = 240

RECEIVER_RIGHT_LABEL_W = 120
RECEIVER_RIGHT_VALUE_W = DOC_WIDTH - (
    RECEIVER_LABEL_W
    + RECEIVER_NAME_W
    + RECEIVER_MID_LABEL_W
    + RECEIVER_MID_VALUE_W
    + RECEIVER_RIGHT_LABEL_W
)
RECEIVER_BLOCK_HEIGHT = 84

# =========================================================
# ☑️ 발신처 블록 폭
# =========================================================
SENDER_LEFT_LABEL_W = 120
SENDER_INNER_LABEL_W = 115

# 사업자등록번호 / 상호 / 대표자 / 주소 / TEL / FAX
SENDER_BIZ_NO_VALUE_W = DOC_WIDTH - SENDER_LEFT_LABEL_W - SENDER_INNER_LABEL_W
SENDER_COMPANY_LEFT_VALUE_W = 480
SENDER_CEO_LABEL_W = 115
SENDER_CEO_VALUE_W = DOC_WIDTH - (
    SENDER_LEFT_LABEL_W
    + SENDER_INNER_LABEL_W
    + SENDER_COMPANY_LEFT_VALUE_W
    + SENDER_CEO_LABEL_W
)
SENDER_ADDR_VALUE_W = DOC_WIDTH - SENDER_LEFT_LABEL_W - SENDER_INNER_LABEL_W
SENDER_TEL_VALUE_W = 480
SENDER_FAX_LABEL_W = 115
SENDER_FAX_VALUE_W = DOC_WIDTH - (
    SENDER_LEFT_LABEL_W
    + SENDER_INNER_LABEL_W
    + SENDER_TEL_VALUE_W
    + SENDER_FAX_LABEL_W
)

SENDER_BLOCK_HEIGHT = FIELD_HEIGHT * 4  # 168

# =========================================================
# ☑️ 발주 정보 영역 폭
# =========================================================
INFO_LEFT_LABEL_W = 120
INFO_LEFT_VALUE_W = 490
INFO_RIGHT_LABEL_W = 120
INFO_RIGHT_VALUE_W = DOC_WIDTH - (
    INFO_LEFT_LABEL_W + INFO_LEFT_VALUE_W + INFO_RIGHT_LABEL_W
)

# =========================================================
# ☑️ 품목 테이블 폭
# =========================================================
COL_NO = 50
COL_LOT = 110
COL_NAME = 150
COL_SPEC = 140
COL_UNIT = 70
COL_QTY = 110
COL_PRICE = 120
COL_SUPPLY = 130
COL_TAX = 120
COL_DEADLINE = 110

ITEM_TOTAL_WIDTH = (
    COL_NO
    + COL_LOT
    + COL_NAME
    + COL_SPEC
    + COL_UNIT
    + COL_QTY
    + COL_PRICE
    + COL_SUPPLY
    + COL_TAX
    + COL_DEADLINE
)

COL_DEADLINE = COL_DEADLINE + (DOC_WIDTH - ITEM_TOTAL_WIDTH)

# =========================================================
# ☑️ 하단 합계 / 참고사항 폭
# =========================================================
SUMMARY_LABEL_W = 120
SUMMARY_VALUE_W = 650
SUMMARY_RIGHT_VALUE_W = DOC_WIDTH - SUMMARY_LABEL_W - SUMMARY_VALUE_W

NOTE_LABEL_W = 120
NOTE_VALUE_W = DOC_WIDTH - NOTE_LABEL_W

CONDITION_LABEL_W = 120
CONDITION_VALUE_W = DOC_WIDTH - CONDITION_LABEL_W


# =========================================================
# ☑️ 숫자 유틸
# =========================================================
def to_int(value):
    # ⭐ 수량/단가/합계 계산 전에 문자열 숫자를 정수로 바꾸는 공통 유틸
    try:
        value = str(value).replace(",", "").strip()
        if value == "":
            return 0
        return int(value)
    except Exception:
        return 0


def format_number(value):
    # ⭐ 계산된 금액(공급가, 세액, 총합계)을 30,000 같은 형식으로 표시
    if value in ("", None):
        return ""
    try:
        return f"{int(value):,}"
    except Exception:
        return str(value)


# =========================================================
# ☑️ 엑셀 스타일 공통 유틸
# =========================================================
# ❌ 엑셀 셀 범위 스타일 공통 함수
# ❌ 테두리 / 정렬 / 폰트 / 배경색 / 값 입력을 한 번에 처리함
# ❌ build_excel_workbook() 안에서 반복 호출되며 엑셀 표 모양을 유지하는 핵심 함수
def apply_range_style(
    ws,
    cell_range,
    value="",
    fill_color=None,
    bold=False,
    font_size=12,
    horizontal="center",
    vertical="center",
):
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    cells = ws[cell_range]
    for row in cells:
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(
                horizontal=horizontal,
                vertical=vertical,
                wrap_text=False,
                shrink_to_fit=True,
            )
            cell.font = Font(
                bold=bold,
                size=font_size,
                name="맑은 고딕",
            )
            if fill_color:
                cell.fill = PatternFill(
                    fill_type="solid",
                    fgColor=fill_color,
                )

    top_left = cells[0][0]
    top_left.value = value


# ❌ 지정한 범위 전체에 테두리와 기본 정렬/폰트를 넣는 함수
# ❌ 품목 표 전체 줄을 그릴 때 사용됨
def set_all_borders(ws, start_row, end_row, start_col, end_col):
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            ws.cell(r, c).border = border
            ws.cell(r, c).alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=False,
                shrink_to_fit=True,
            )
            ws.cell(r, c).font = Font(name="맑은 고딕", size=11)


# =========================================================
# ☑️ 공통 셀
# =========================================================
def box_label(text, width=None, height=FIELD_HEIGHT, bold=False):
    return ft.Container(
        width=width,
        height=height,
        bgcolor=HEADER_BG,
        alignment=ft.Alignment(0, 0),
        border=ft.Border.all(1, BORDER_COLOR),
        content=ft.Text(
            text,
            size=13,
            color=TEXT_COLOR,
            weight=ft.FontWeight.W_700 if bold else ft.FontWeight.W_500,
            text_align=ft.TextAlign.CENTER,
        ),
    )


def box_input(control, width=None, height=FIELD_HEIGHT, bgcolor=WHITE):
    return ft.Container(
        width=width,
        height=height,
        border=ft.Border.all(1, BORDER_COLOR),
        bgcolor=bgcolor,
        alignment=ft.Alignment(0, 0),
        content=control,
    )


def build_textfield(
    value="",
    text_align=ft.TextAlign.LEFT,
    read_only=False,
    on_change=None,
    dense=True,
):
    return ft.TextField(
        value=value,
        border=ft.InputBorder.NONE,
        filled=False,
        text_size=13,
        cursor_height=18,
        content_padding=ft.Padding(left=8, right=8, top=10, bottom=10),
        text_align=text_align,
        read_only=read_only,
        on_change=on_change,
        dense=dense,
    )


# =========================================================
# ☑️ 발주서 모달 클래스
# =========================================================
class PurchaseOrderDialog:
    def __init__(self, page: ft.Page):
        self.page = page

        # -------------------------------------------------
        # ☑️ 상단 결재란
        # -------------------------------------------------
        self.approval_1 = build_textfield(text_align=ft.TextAlign.CENTER)
        self.approval_2 = build_textfield(text_align=ft.TextAlign.CENTER)
        self.approval_3 = build_textfield(text_align=ft.TextAlign.CENTER)

        # -------------------------------------------------
        # ☑️ 수신처
        # -------------------------------------------------
        self.receiver_name = build_textfield()
        self.receiver_tel = build_textfield()
        self.receiver_fax = build_textfield()
        self.receiver_person = build_textfield()

        # -------------------------------------------------
        # ☑️ 발신처
        # -------------------------------------------------
        self.sender_business_no = build_textfield()
        self.sender_company_name = build_textfield()
        self.sender_ceo = build_textfield()
        self.sender_address = build_textfield()
        self.sender_tel = build_textfield()
        self.sender_fax = build_textfield()

        # -------------------------------------------------
        # ☑️ 발주 정보
        # -------------------------------------------------
        self.order_date = build_textfield()
        self.order_no = build_textfield()
        self.contract_no = build_textfield()

        self.delivery_place = build_textfield()
        self.request_dept = build_textfield()
        self.order_manager = build_textfield()

        # -------------------------------------------------
        # ☑️ 하단 참고사항 / 발주조건 / 합계
        # -------------------------------------------------
        self.total_amount_text = build_textfield(
            text_align=ft.TextAlign.RIGHT,
            read_only=True,
        )
        self.total_amount_krw = build_textfield(read_only=True)

        self.note_text = build_textfield()
        self.payment_condition = build_textfield()

        # -------------------------------------------------
        # ☑️ 품목 행
        # -------------------------------------------------
        self.item_rows = []
        for i in range(20):
            self.item_rows.append(self.build_item_row(i + 1))

        # -------------------------------------------------
        # ☑️ 다이얼로그
        # -------------------------------------------------
        self.dialog = ft.AlertDialog(
            modal=True,
            inset_padding=10,
            bgcolor=ft.Colors.TRANSPARENT,
            content_padding=0,
            shape=ft.RoundedRectangleBorder(radius=12),
            content=self.build_dialog_content(),
        )

    # =====================================================
    # ☑️ 숫자를 한글 금액으로 간단 변환
    # =====================================================
    def number_to_korean(self, value: int) -> str:
        if value == 0:
            return "영원"

        units = ["", "만", "억", "조"]
        small_units = ["", "십", "백", "천"]
        nums = ["", "일", "이", "삼", "사", "오", "육", "칠", "팔", "구"]

        result_parts = []
        unit_index = 0

        while value > 0:
            part = value % 10000
            value //= 10000

            if part > 0:
                part_str = ""
                digits = list(map(int, str(part)))
                length = len(digits)

                for i, n in enumerate(digits):
                    if n == 0:
                        continue
                    small_unit = small_units[length - i - 1]

                    if n == 1 and small_unit != "":
                        part_str += small_unit
                    else:
                        part_str += nums[n] + small_unit

                part_str += units[unit_index]
                result_parts.insert(0, part_str)

            unit_index += 1

        return "".join(result_parts) + "원"

    # =====================================================
    # ☑️ 품목 한 줄 생성
    # =====================================================
    def build_item_row(self, no: int):
        row_data = {}

        def on_amount_change(e=None):
            qty = to_int(row_data["qty"].value)
            price = to_int(row_data["unit_price"].value)

            supply = qty * price
            tax = int(supply * 0.1)

            row_data["supply_amount"].value = format_number(supply) if supply else ""
            row_data["tax_amount"].value = format_number(tax) if tax else ""

            self.update_summary()
            self.page.update()

        row_data["no"] = no
        row_data["lot_no"] = build_textfield(text_align=ft.TextAlign.CENTER)
        row_data["product_name"] = build_textfield()
        row_data["spec"] = build_textfield()
        row_data["unit"] = build_textfield(text_align=ft.TextAlign.CENTER)

        row_data["qty"] = build_textfield(
            text_align=ft.TextAlign.RIGHT,
            on_change=on_amount_change,
        )
        row_data["unit_price"] = build_textfield(
            text_align=ft.TextAlign.RIGHT,
            on_change=on_amount_change,
        )
        row_data["supply_amount"] = build_textfield(
            text_align=ft.TextAlign.RIGHT,
            read_only=True,
        )
        row_data["tax_amount"] = build_textfield(
            text_align=ft.TextAlign.RIGHT,
            read_only=True,
        )
        row_data["delivery_deadline"] = build_textfield(text_align=ft.TextAlign.CENTER)

        return row_data

    # =====================================================
    # ☑️ 합계 업데이트
    # =====================================================
    # ❌ 발주서 품목행의 공급가/세액을 모두 합산해서 총합계를 만드는 핵심 함수
    # ❌ total_amount_text, total_amount_krw 값을 갱신하며 엑셀 생성 직전에도 호출됨
    def update_summary(self):
        total_supply = 0
        total_tax = 0

        for row in self.item_rows:
            total_supply += to_int(row["supply_amount"].value)
            total_tax += to_int(row["tax_amount"].value)

        grand_total = total_supply + total_tax

        self.total_amount_text.value = format_number(grand_total) if grand_total else ""
        self.total_amount_krw.value = (
            self.number_to_korean(grand_total) if grand_total else ""
        )

    # =====================================================
    # ☑️ 제목 + 결재란
    # =====================================================
    def build_title_area(self):
        return ft.Container(
            width=DOC_WIDTH,
            content=ft.Row(
                spacing=0,
                controls=[
                    ft.Container(
                        width=TITLE_LEFT_WIDTH,
                        height=TITLE_HEIGHT,
                        border=ft.Border.all(1, BORDER_COLOR),
                        alignment=ft.Alignment(0, 0),
                        content=ft.Text(
                            "발 주 서",
                            size=28,
                            weight=ft.FontWeight.W_700,
                            color=TEXT_COLOR,
                        ),
                    ),
                    ft.Container(
                        width=TITLE_RIGHT_WIDTH,
                        height=TITLE_HEIGHT,
                        content=ft.Column(
                            spacing=0,
                            controls=[
                                ft.Row(
                                    spacing=0,
                                    controls=[
                                        box_label("담당", width=APPROVAL_COL_WIDTH),
                                        box_input(self.approval_1, width=APPROVAL_COL_WIDTH),
                                        box_input(self.approval_2, width=APPROVAL_COL_WIDTH),
                                    ],
                                ),
                                ft.Row(
                                    spacing=0,
                                    controls=[
                                        box_input(
                                            build_textfield(),
                                            width=APPROVAL_COL_WIDTH,
                                            height=108,
                                        ),
                                        box_input(
                                            build_textfield(),
                                            width=APPROVAL_COL_WIDTH,
                                            height=108,
                                        ),
                                        box_input(
                                            build_textfield(),
                                            width=APPROVAL_COL_WIDTH,
                                            height=108,
                                        ),
                                    ],
                                ),
                            ],
                        ),
                    ),
                ],
            ),
        )

    # =====================================================
    # ☑️ 수신처 블록
    # =====================================================
    def build_receiver_section(self):
        return ft.Container(
            width=DOC_WIDTH,
            content=ft.Row(
                spacing=0,
                controls=[
                    box_label("수신처", width=RECEIVER_LABEL_W, height=RECEIVER_BLOCK_HEIGHT),
                    box_input(self.receiver_name, width=RECEIVER_NAME_W, height=RECEIVER_BLOCK_HEIGHT),
                    ft.Column(
                        spacing=0,
                        controls=[
                            ft.Row(
                                spacing=0,
                                controls=[
                                    box_label("TEL", width=RECEIVER_MID_LABEL_W, bold=True),
                                    box_input(self.receiver_tel, width=RECEIVER_MID_VALUE_W),
                                ],
                            ),
                            ft.Row(
                                spacing=0,
                                controls=[
                                    box_label("FAX", width=RECEIVER_MID_LABEL_W, bold=True),
                                    box_input(self.receiver_fax, width=RECEIVER_MID_VALUE_W),
                                ],
                            ),
                        ],
                    ),
                    ft.Row(
                        spacing=0,
                        controls=[
                            box_label("담당자", width=RECEIVER_RIGHT_LABEL_W, height=RECEIVER_BLOCK_HEIGHT),
                            box_input(
                                self.receiver_person,
                                width=RECEIVER_RIGHT_VALUE_W,
                                height=RECEIVER_BLOCK_HEIGHT,
                            ),
                        ],
                    ),
                ],
            ),
        )

    # =====================================================
    # ☑️ 발신처 블록
    # =====================================================
    def build_sender_section(self):
        return ft.Container(
            width=DOC_WIDTH,
            content=ft.Row(
                spacing=0,
                controls=[
                    box_label("발신처", width=SENDER_LEFT_LABEL_W, height=SENDER_BLOCK_HEIGHT),
                    ft.Column(
                        spacing=0,
                        controls=[
                            ft.Row(
                                spacing=0,
                                controls=[
                                    box_label("사업자 등록번호", width=SENDER_INNER_LABEL_W),
                                    box_input(self.sender_business_no, width=SENDER_BIZ_NO_VALUE_W),
                                ],
                            ),
                            ft.Row(
                                spacing=0,
                                controls=[
                                    box_label("상호", width=SENDER_INNER_LABEL_W),
                                    box_input(self.sender_company_name, width=SENDER_COMPANY_LEFT_VALUE_W),
                                    box_label("대표자", width=SENDER_CEO_LABEL_W),
                                    box_input(self.sender_ceo, width=SENDER_CEO_VALUE_W),
                                ],
                            ),
                            ft.Row(
                                spacing=0,
                                controls=[
                                    box_label("주소", width=SENDER_INNER_LABEL_W),
                                    box_input(self.sender_address, width=SENDER_ADDR_VALUE_W),
                                ],
                            ),
                            ft.Row(
                                spacing=0,
                                controls=[
                                    box_label("TEL", width=SENDER_INNER_LABEL_W, bold=True),
                                    box_input(self.sender_tel, width=SENDER_TEL_VALUE_W),
                                    box_label("FAX", width=SENDER_FAX_LABEL_W, bold=True),
                                    box_input(self.sender_fax, width=SENDER_FAX_VALUE_W),
                                ],
                            ),
                        ],
                    ),
                ],
            ),
        )

    # =====================================================
    # ☑️ 발주 정보 2열 공통 행
    # =====================================================
    def build_info_row(self, left_title, left_control, right_title, right_control):
        return ft.Container(
            width=DOC_WIDTH,
            content=ft.Row(
                spacing=0,
                controls=[
                    box_label(left_title, width=INFO_LEFT_LABEL_W),
                    box_input(left_control, width=INFO_LEFT_VALUE_W),
                    box_label(right_title, width=INFO_RIGHT_LABEL_W),
                    box_input(right_control, width=INFO_RIGHT_VALUE_W),
                ],
            ),
        )

    # =====================================================
    # ☑️ 품목 헤더
    # =====================================================
    def build_item_header(self):
        return ft.Container(
            width=DOC_WIDTH,
            content=ft.Row(
                spacing=0,
                controls=[
                    box_label("no", width=COL_NO, height=44, bold=True),
                    box_label("LOT NO", width=COL_LOT, height=44, bold=True),
                    box_label("품명", width=COL_NAME, height=44, bold=True),
                    box_label("규격/사양", width=COL_SPEC, height=44, bold=True),
                    box_label("단위", width=COL_UNIT, height=44, bold=True),
                    box_label("수량", width=COL_QTY, height=44, bold=True),
                    box_label("단가", width=COL_PRICE, height=44, bold=True),
                    box_label("공급가계", width=COL_SUPPLY, height=44, bold=True),
                    box_label("세액", width=COL_TAX, height=44, bold=True),
                    box_label("납품기한", width=COL_DEADLINE, height=44, bold=True),
                ],
            ),
        )

    # =====================================================
    # ☑️ 품목 한 줄 UI
    # =====================================================
    def build_item_line(self, row_data):
        return ft.Container(
            width=DOC_WIDTH,
            content=ft.Row(
                spacing=0,
                controls=[
                    box_input(
                        build_textfield(
                            value=str(row_data["no"]),
                            text_align=ft.TextAlign.CENTER,
                            read_only=True,
                        ),
                        width=COL_NO,
                        height=ROW_HEIGHT,
                    ),
                    box_input(row_data["lot_no"], width=COL_LOT, height=ROW_HEIGHT),
                    box_input(row_data["product_name"], width=COL_NAME, height=ROW_HEIGHT),
                    box_input(row_data["spec"], width=COL_SPEC, height=ROW_HEIGHT),
                    box_input(row_data["unit"], width=COL_UNIT, height=ROW_HEIGHT),
                    box_input(row_data["qty"], width=COL_QTY, height=ROW_HEIGHT),
                    box_input(row_data["unit_price"], width=COL_PRICE, height=ROW_HEIGHT),
                    box_input(row_data["supply_amount"], width=COL_SUPPLY, height=ROW_HEIGHT),
                    box_input(row_data["tax_amount"], width=COL_TAX, height=ROW_HEIGHT),
                    box_input(row_data["delivery_deadline"], width=COL_DEADLINE, height=ROW_HEIGHT),
                ],
            ),
        )

    # =====================================================
    # ☑️ 품목 입력 영역
    # =====================================================
    def build_items_section(self):
        return ft.Container(
            width=DOC_WIDTH,
            content=ft.Column(
                spacing=0,
                controls=[self.build_item_header()]
                + [self.build_item_line(row_data) for row_data in self.item_rows],
            ),
        )

    # =====================================================
    # ☑️ 하단 합계 영역
    # =====================================================
    def build_summary_section(self):
        return ft.Container(
            width=DOC_WIDTH,
            content=ft.Row(
                spacing=0,
                controls=[
                    box_label("총합계", width=SUMMARY_LABEL_W, height=44, bold=True),
                    box_input(self.total_amount_krw, width=SUMMARY_VALUE_W, height=44),
                    box_input(
                        self.total_amount_text,
                        width=SUMMARY_RIGHT_VALUE_W,
                        height=44,
                        bgcolor=WHITE,
                    ),
                ],
            ),
        )

    # =====================================================
    # ☑️ 참고사항
    # =====================================================
    def build_note_section(self):
        return ft.Container(
            width=DOC_WIDTH,
            content=ft.Row(
                spacing=0,
                controls=[
                    box_label("참고사항", width=NOTE_LABEL_W, height=56),
                    box_input(self.note_text, width=NOTE_VALUE_W, height=56),
                ],
            ),
        )

    # =====================================================
    # ☑️ 발주조건
    # =====================================================
    def build_condition_section(self):
        return ft.Container(
            width=DOC_WIDTH,
            content=ft.Row(
                spacing=0,
                controls=[
                    box_label("발주조건", width=CONDITION_LABEL_W, height=56),
                    box_input(self.payment_condition, width=CONDITION_VALUE_W, height=56),
                ],
            ),
        )

    # =====================================================
    # ☑️ 저장 데이터 수집
    # =====================================================
    # ❌ 화면 입력값들을 dict 형태의 저장 데이터로 모으는 핵심 함수
    # ❌ 엑셀 생성 전에 build_excel_workbook()가 이 데이터를 가져다 씀
    def collect_data(self):
        items = []

        for row in self.item_rows:
            item = {
                "no": row["no"],
                "lot_no": row["lot_no"].value,
                "product_name": row["product_name"].value,
                "spec": row["spec"].value,
                "unit": row["unit"].value,
                "qty": row["qty"].value,
                "unit_price": row["unit_price"].value,
                "supply_amount": row["supply_amount"].value,
                "tax_amount": row["tax_amount"].value,
                "delivery_deadline": row["delivery_deadline"].value,
            }

            has_value = any(str(v).strip() != "" for k, v in item.items() if k != "no")
            if has_value:
                items.append(item)

        return {
            "approval_1": self.approval_1.value,
            "approval_2": self.approval_2.value,
            "approval_3": self.approval_3.value,
            "receiver_name": self.receiver_name.value,
            "receiver_tel": self.receiver_tel.value,
            "receiver_fax": self.receiver_fax.value,
            "receiver_person": self.receiver_person.value,
            "sender_business_no": self.sender_business_no.value,
            "sender_company_name": self.sender_company_name.value,
            "sender_ceo": self.sender_ceo.value,
            "sender_address": self.sender_address.value,
            "sender_tel": self.sender_tel.value,
            "sender_fax": self.sender_fax.value,
            "order_date": self.order_date.value,
            "order_no": self.order_no.value,
            "contract_no": self.contract_no.value,
            "delivery_place": self.delivery_place.value,
            "request_dept": self.request_dept.value,
            "order_manager": self.order_manager.value,
            "total_amount_text": self.total_amount_text.value,
            "total_amount_krw": self.total_amount_krw.value,
            "note_text": self.note_text.value,
            "payment_condition": self.payment_condition.value,
            "items": items,
        }

    # =====================================================
    # ☑️ 엑셀 워크북 생성
    # =====================================================
    # ❌ 발주서 화면 구조를 엑셀 워크북으로 그대로 다시 만드는 핵심 함수
    # ❌ 컬럼 폭 / 행 높이 / 병합 / 제목 / 수신처 / 발신처 / 발주정보 / 품목표 / 총합계 / 참고사항 / 발주조건이 전부 여기 있음
    def build_excel_workbook(self):
        self.update_summary()
        data = self.collect_data()

        wb = Workbook()
        ws = wb.active
        ws.title = "발주서"

        # -------------------------------------------------
        # ☑️ 컬럼 폭 설정
        # -------------------------------------------------
        column_widths = {
            "A": 12,
            "B": 18,
            "C": 18,
            "D": 16,
            "E": 12,
            "F": 14,
            "G": 12,
            "H": 14,
            "I": 12,
            "J": 16,
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # -------------------------------------------------
        # ☑️ 행 높이 설정
        # -------------------------------------------------
        for row_idx in range(1, 60):
            ws.row_dimensions[row_idx].height = 26

        ws.row_dimensions[1].height = 34
        ws.row_dimensions[2].height = 34
        ws.row_dimensions[3].height = 34
        ws.row_dimensions[4].height = 34
        ws.row_dimensions[18].height = 30
        ws.row_dimensions[39].height = 32
        ws.row_dimensions[40].height = 36
        ws.row_dimensions[41].height = 42

        # -------------------------------------------------
        # ☑️ 머지
        # -------------------------------------------------
        ws.merge_cells("A1:G4")

        # 오른쪽 결재란
        ws.merge_cells("H2:H4")
        ws.merge_cells("I2:I4")
        ws.merge_cells("J2:J4")

        # 수신처
        ws.merge_cells("A5:A6")
        ws.merge_cells("B5:D6")
        ws.merge_cells("F5:G5")
        ws.merge_cells("F6:G6")
        ws.merge_cells("H5:H6")
        ws.merge_cells("I5:J6")

        # 발신처
        ws.merge_cells("A7:A10")
        ws.merge_cells("C7:J7")
        ws.merge_cells("C8:F8")
        ws.merge_cells("H8:J8")
        ws.merge_cells("C9:J9")
        ws.merge_cells("C10:F10")
        ws.merge_cells("H10:J10")

        # 발주 정보
        ws.merge_cells("B11:E11")
        ws.merge_cells("G11:J11")
        ws.merge_cells("B12:E12")
        ws.merge_cells("G12:J12")
        ws.merge_cells("B13:E13")
        ws.merge_cells("G13:J13")

        # 하단
        ws.merge_cells("B40:F40")
        ws.merge_cells("G40:J40")
        ws.merge_cells("B41:J41")
        ws.merge_cells("B42:J42")

        # -------------------------------------------------
        # ☑️ 공통 색상
        # -------------------------------------------------
        header_fill = "F2F2F2"

        # -------------------------------------------------
        # ☑️ 제목 영역
        # -------------------------------------------------
        apply_range_style(
            ws,
            "A1:G4",
            value="발 주 서",
            bold=True,
            font_size=24,
        )

        apply_range_style(ws, "H1:H1", value="담당", fill_color=header_fill, bold=True)
        apply_range_style(ws, "I1:I1", value="", fill_color=header_fill, bold=True)
        apply_range_style(ws, "J1:J1", value="", fill_color=header_fill, bold=True)

        apply_range_style(ws, "H2:H4", value=data["approval_1"])
        apply_range_style(ws, "I2:I4", value=data["approval_2"])
        apply_range_style(ws, "J2:J4", value=data["approval_3"])

        # -------------------------------------------------
        # ☑️ 수신처
        # -------------------------------------------------
        apply_range_style(ws, "A5:A6", value="수신처", fill_color=header_fill, bold=True)
        apply_range_style(ws, "B5:D6", value=data["receiver_name"], horizontal="left")

        apply_range_style(ws, "E5:E5", value="TEL", fill_color=header_fill, bold=True)
        apply_range_style(ws, "F5:G5", value=data["receiver_tel"], horizontal="left")

        apply_range_style(ws, "E6:E6", value="FAX", fill_color=header_fill, bold=True)
        apply_range_style(ws, "F6:G6", value=data["receiver_fax"], horizontal="left")

        apply_range_style(ws, "H5:H6", value="담당자", fill_color=header_fill, bold=True)
        apply_range_style(ws, "I5:J6", value=data["receiver_person"], horizontal="left")

        # -------------------------------------------------
        # ☑️ 발신처
        # -------------------------------------------------
        apply_range_style(ws, "A7:A10", value="발신처", fill_color=header_fill, bold=True)

        apply_range_style(ws, "B7:B7", value="사업자 등록번호", fill_color=header_fill, bold=True)
        apply_range_style(ws, "C7:J7", value=data["sender_business_no"], horizontal="left")

        apply_range_style(ws, "B8:B8", value="상호", fill_color=header_fill, bold=True)
        apply_range_style(ws, "C8:F8", value=data["sender_company_name"], horizontal="left")
        apply_range_style(ws, "G8:G8", value="대표자", fill_color=header_fill, bold=True)
        apply_range_style(ws, "H8:J8", value=data["sender_ceo"], horizontal="left")

        apply_range_style(ws, "B9:B9", value="주소", fill_color=header_fill, bold=True)
        apply_range_style(ws, "C9:J9", value=data["sender_address"], horizontal="left")

        apply_range_style(ws, "B10:B10", value="TEL", fill_color=header_fill, bold=True)
        apply_range_style(ws, "C10:F10", value=data["sender_tel"], horizontal="left")
        apply_range_style(ws, "G10:G10", value="FAX", fill_color=header_fill, bold=True)
        apply_range_style(ws, "H10:J10", value=data["sender_fax"], horizontal="left")

        # -------------------------------------------------
        # ☑️ 발주 정보
        # -------------------------------------------------
        apply_range_style(ws, "A11:A11", value="발주일자", fill_color=header_fill, bold=True)
        apply_range_style(ws, "B11:E11", value=data["order_date"], horizontal="left")
        apply_range_style(ws, "F11:F11", value="납품장소", fill_color=header_fill, bold=True)
        apply_range_style(ws, "G11:J11", value=data["delivery_place"], horizontal="left")

        apply_range_style(ws, "A12:A12", value="발주번호", fill_color=header_fill, bold=True)
        apply_range_style(ws, "B12:E12", value=data["order_no"], horizontal="left")
        apply_range_style(ws, "F12:F12", value="요구부서", fill_color=header_fill, bold=True)
        apply_range_style(ws, "G12:J12", value=data["request_dept"], horizontal="left")

        apply_range_style(ws, "A13:A13", value="계약번호", fill_color=header_fill, bold=True)
        apply_range_style(ws, "B13:E13", value=data["contract_no"], horizontal="left")
        apply_range_style(ws, "F13:F13", value="발주담당자", fill_color=header_fill, bold=True)
        apply_range_style(ws, "G13:J13", value=data["order_manager"], horizontal="left")

        # -------------------------------------------------
        # ☑️ 품목 헤더
        # -------------------------------------------------
        item_header_row = 14
        headers = [
            "no",
            "LOT NO",
            "품명",
            "규격/사양",
            "단위",
            "수량",
            "단가",
            "공급가계",
            "세액",
            "납품기한",
        ]

        for idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=item_header_row, column=idx)
            cell.value = header
            cell.fill = PatternFill(fill_type="solid", fgColor=header_fill)
            cell.font = Font(bold=True, size=11, name="맑은 고딕")
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=False,
                shrink_to_fit=True,
            )

        set_all_borders(ws, item_header_row, item_header_row, 1, 10)

        # -------------------------------------------------
        # ☑️ 품목 데이터
        # -------------------------------------------------
        start_row = 15
        max_item_rows = 25
        items = data["items"]

        if not items:
            for i in range(max_item_rows):
                row_no = start_row + i
                ws.cell(row=row_no, column=1).value = i + 1
            set_all_borders(ws, start_row, start_row + max_item_rows - 1, 1, 10)
        else:
            for i in range(max_item_rows):
                row_no = start_row + i
                ws.cell(row=row_no, column=1).value = i + 1

            for i, item in enumerate(items[:max_item_rows], start=0):
                row_no = start_row + i
                ws.cell(row=row_no, column=1).value = item["no"]
                ws.cell(row=row_no, column=2).value = item["lot_no"]
                ws.cell(row=row_no, column=3).value = item["product_name"]
                ws.cell(row=row_no, column=4).value = item["spec"]
                ws.cell(row=row_no, column=5).value = item["unit"]
                ws.cell(row=row_no, column=6).value = item["qty"]
                ws.cell(row=row_no, column=7).value = item["unit_price"]
                ws.cell(row=row_no, column=8).value = item["supply_amount"]
                ws.cell(row=row_no, column=9).value = item["tax_amount"]
                ws.cell(row=row_no, column=10).value = item["delivery_deadline"]

            set_all_borders(ws, start_row, start_row + max_item_rows - 1, 1, 10)

        # -------------------------------------------------
        # ☑️ 하단 합계 / 참고사항 / 발주조건
        # -------------------------------------------------
        apply_range_style(ws, "A40:A40", value="총합계", fill_color=header_fill, bold=True)
        apply_range_style(ws, "B40:F40", value=data["total_amount_krw"], horizontal="left")
        apply_range_style(ws, "G40:J40", value=data["total_amount_text"], horizontal="right")

        apply_range_style(ws, "A41:A41", value="참고사항", fill_color=header_fill, bold=True)
        apply_range_style(ws, "B41:J41", value=data["note_text"], horizontal="left")

        apply_range_style(ws, "A42:A42", value="발주조건", fill_color=header_fill, bold=True)
        apply_range_style(ws, "B42:J42", value=data["payment_condition"], horizontal="left")

        # -------------------------------------------------
        # ☑️ 시트 옵션
        # -------------------------------------------------
        ws.freeze_panes = None
        ws.sheet_view.showGridLines = False

        return wb

    # =====================================================
    # ☑️ 엑셀 내보내기
    # =====================================================
    # ❌ build_excel_workbook()로 만든 워크북을 실제 .xlsx 파일로 저장하는 함수
    # ❌ 저장 경로 결정, wb.save(), 저장 완료 메시지 출력이 전부 여기서 처리됨
    def export_to_excel(self, e):
        try:
            wb = self.build_excel_workbook()
            base_dir = os.path.dirname(os.path.abspath(__file__))
            save_path = os.path.join(base_dir, "purchase_order.xlsx")
            wb.save(save_path)

            print("엑셀 저장 완료:", save_path)

            self.page.snack_bar = ft.SnackBar(
                content=ft.Text(f"엑셀 파일이 저장되었습니다. ({save_path})"),
                open=True,
            )
            self.page.update()

        except Exception as ex:
            print("엑셀 저장 오류:", ex)
            self.page.snack_bar = ft.SnackBar(
                content=ft.Text(f"엑셀 저장 중 오류가 발생했습니다: {ex}"),
                open=True,
            )
            self.page.update()

    # =====================================================
    # ☑️ 저장 버튼
    # =====================================================
    def save_data(self, e):
        self.update_summary()
        data = self.collect_data()

        print("===== 발주서 저장 데이터 =====")
        print(data)

        self.page.snack_bar = ft.SnackBar(
            content=ft.Text("발주서가 저장되었습니다."),
            open=True,
        )
        self.page.update()

    # =====================================================
    # ☑️ 닫기 버튼
    # =====================================================
    def close_dialog(self, e=None):
        self.page.pop_dialog()
        self.page.update()

    # =====================================================
    # ☑️ 전체 콘텐츠
    # =====================================================
    def build_dialog_content(self):
        return ft.Container(
            width=DIALOG_WIDTH,
            height=DIALOG_HEIGHT,
            bgcolor=WHITE,
            border_radius=12,
            padding=20,
            content=ft.Column(
                spacing=0,
                controls=[
                    ft.Row(
                        alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
                        vertical_alignment=ft.CrossAxisAlignment.CENTER,
                        controls=[
                            ft.Text(
                                "발주서 입력",
                                size=20,
                                weight=ft.FontWeight.W_700,
                                color=TEXT_COLOR,
                            ),
                            ft.IconButton(
                                icon=ft.Icons.CLOSE,
                                on_click=self.close_dialog,
                            ),
                        ],
                    ),
                    ft.Container(height=16),
                    ft.Container(
                        expand=True,
                        content=ft.Column(
                            scroll=ft.ScrollMode.AUTO,
                            spacing=0,
                            horizontal_alignment=ft.CrossAxisAlignment.CENTER,
                            controls=[
                                self.build_title_area(),
                                self.build_receiver_section(),
                                self.build_sender_section(),
                                self.build_info_row("발주일자", self.order_date, "납품장소", self.delivery_place),
                                self.build_info_row("발주번호", self.order_no, "요구부서", self.request_dept),
                                self.build_info_row("계약번호", self.contract_no, "발주담당자", self.order_manager),
                                self.build_items_section(),
                                self.build_summary_section(),
                                self.build_note_section(),
                                self.build_condition_section(),
                            ],
                        ),
                    ),
                    ft.Container(height=16),
                    ft.Row(
                        alignment=ft.MainAxisAlignment.END,
                        controls=[
                            ft.ElevatedButton(
                                "닫기",
                                height=42,
                                style=ft.ButtonStyle(
                                    shape=ft.RoundedRectangleBorder(radius=8),
                                ),
                                on_click=self.close_dialog,
                            ),
                            ft.ElevatedButton(
                                "엑셀 내보내기",
                                height=42,
                                style=ft.ButtonStyle(
                                    bgcolor="#16A34A",
                                    color=ft.Colors.WHITE,
                                    shape=ft.RoundedRectangleBorder(radius=8),
                                ),
                                on_click=self.export_to_excel,
                            ),
                            ft.ElevatedButton(
                                "저장",
                                height=42,
                                style=ft.ButtonStyle(
                                    bgcolor="#2563EB",
                                    color=ft.Colors.WHITE,
                                    shape=ft.RoundedRectangleBorder(radius=8),
                                ),
                                on_click=self.save_data,
                            ),
                        ],
                    ),
                ],
            ),
        )

    # =====================================================
    # ☑️ 모달 열기
    # =====================================================
    def open(self, e=None):
        self.page.show_dialog(self.dialog)
        self.page.update()


# =========================================================
# ☑️ 메인 실행부
# =========================================================
def main(page: ft.Page):
    page.title = "발주서 모달 예제"
    page.bgcolor = PAGE_BG
    page.padding = 30
    page.scroll = ft.ScrollMode.AUTO

    popup = PurchaseOrderDialog(page)

    purchase_order = ft.Column(
        alignment=ft.MainAxisAlignment.CENTER,
        horizontal_alignment=ft.CrossAxisAlignment.CENTER,
        controls=[
            ft.Text(
                "발주서 모달창 예제",
                size=26,
                weight=ft.FontWeight.W_700,
            ),
            ft.Container(height=20),
            ft.ElevatedButton(
                "발주서 열기",
                width=220,
                height=50,
                style=ft.ButtonStyle(
                    bgcolor="#111827",
                    color=ft.Colors.WHITE,
                    shape=ft.RoundedRectangleBorder(radius=10),
                ),
                on_click=popup.open,
            ),
        ],
    )

    page.add(purchase_order)


ft.run(main)